#data driven with excel

import openpyxl
excelfilepath = "F:\sreaddatafromexcelselenium.xlsx"  # location of excel file.
varworkbook = openpyxl.load_workbook(excelfilepath)

# excel file may have multiple sheets . SO we first need to select the sheet where we want to pergorm actions.
sheet = varworkbook.active # 1. if we have only one sheet in excel file then we use this.
# sheet = varworkbook.get_sheet_by_name("Sheet2") # 2. If there are multiple sheets in excel file then we select sheet by this method.
 # line 9 ko "Sheet1" yesto excel ko bottom ma deakha hunxa kun sheet ma kaam garirako ho vanera excel ma.
# we have only one sheet in excel file so we use 1. option.

noofrows = sheet.max_row
noofcolumn  = sheet.max_column
print(noofrows) #11
print(noofcolumn) #4

#reading data from excel

for r in range(1, noofrows + 1):#(1, noofrows) default takes n-1 rows .ie.11-1 in this case file. so we need to add 1
    for c in range(1,   noofcolumn + 1):
        print(sheet.cell(row=r, column=c).value, end="   ")

    print()

