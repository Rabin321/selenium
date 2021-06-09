import openpyxl

excelfilepath = "F:\writeintoexcelfileselenium.xlsx"

varworkbook = openpyxl.load_workbook(excelfilepath)
sheet = varworkbook.active
for r in range(1, 6):
    for c in range(1, 5):
        sheet.cell(row=r, column=c).value = " Name"
        sheet.cell(row=r, column=c).value = "Ram"

varworkbook.save(excelfilepath)


